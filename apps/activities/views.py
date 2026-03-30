
"""
活动视图 - 校园打卡平台
"""
import csv
import io
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, F, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from rest_framework import permissions, viewsets
from rest_framework.exceptions import PermissionDenied

from apps.checkins.models import CheckIn
from apps.checkins.utils import award_points, calculate_continuous_days
from apps.social.models import Message, Moment

from .forms import ActivityApplicationForm, ActivityCommentForm, ActivityForm
from .models import Activity, ActivityApplication, ActivityComment, ActivityRegistration, Category
from .serializers import ActivityRegistrationSerializer, ActivitySerializer


def _can_manage_activity(user, activity):
    if not user or not user.is_authenticated:
        return False
    return activity.can_edit(user) or activity.can_close(user)


def _management_permission_denied(request, activity, message='您没有权限管理此活动。'):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'message': message}, status=403)
    messages.error(request, message)
    return redirect('activities:detail', pk=activity.pk)


def _export_permission_denied(request, activity, message='您没有权限导出此活动数据。'):
    messages.error(request, message)
    return redirect('activities:detail', pk=activity.pk)


def _require_platform_admin(request):
    if getattr(request.user, 'role', '') != 'admin':
        messages.error(request, '只有平台管理员可以处理活动申请。')
        return False
    return True


def _filtered_get_params(querydict, exclude_keys=None):
    exclude_keys = exclude_keys or []
    data = querydict.copy()
    for key in exclude_keys:
        if key in data:
            del data[key]
    return data.urlencode()


def _get_participant_queryset(activity, keyword='', status=''):
    qs = ActivityRegistration.objects.filter(activity=activity).select_related('user').order_by('-registered_at')
    if keyword:
        qs = qs.filter(
            Q(user__username__icontains=keyword) |
            Q(user__real_name__icontains=keyword) |
            Q(user__department__icontains=keyword)
        )
    if status:
        qs = qs.filter(status=status)
    return qs


def _get_checkin_queryset(activity, keyword='', status=''):
    qs = (
        CheckIn.objects.filter(activity=activity)
        .select_related('user', 'reviewed_by', 'registration')
        .prefetch_related('photos')
        .order_by('-created_at')
    )
    if keyword:
        qs = qs.filter(
            Q(user__username__icontains=keyword) |
            Q(user__real_name__icontains=keyword) |
            Q(remark__icontains=keyword) |
            Q(review_note__icontains=keyword) |
            Q(location_name__icontains=keyword)
        )
    if status:
        qs = qs.filter(status=status)
    return qs


def _get_moment_queryset(activity, keyword=''):
    qs = (
        Moment.objects.filter(activity=activity)
        .select_related('user')
        .prefetch_related('images')
        .annotate(
            like_count=Count('likes', distinct=True),
            comment_count=Count('comments', distinct=True),
        )
        .order_by('-created_at')
    )
    if keyword:
        qs = qs.filter(Q(user__username__icontains=keyword) | Q(content__icontains=keyword))
    return qs


def _get_comment_queryset(activity, keyword=''):
    qs = activity.comments.select_related('user').order_by('-created_at')
    if keyword:
        qs = qs.filter(Q(user__username__icontains=keyword) | Q(content__icontains=keyword))
    return qs


def _build_activity_management_data(activity, request):
    participant_keyword = request.GET.get('participant_q', '').strip()
    participant_status = request.GET.get('participant_status', '').strip()
    participant_page_number = request.GET.get('participant_page')

    checkin_keyword = request.GET.get('checkin_q', '').strip()
    checkin_status = request.GET.get('checkin_status', '').strip()
    checkin_page_number = request.GET.get('checkin_page')

    moment_keyword = request.GET.get('moment_q', '').strip()
    moment_page_number = request.GET.get('moment_page')

    comment_keyword = request.GET.get('comment_q', '').strip()
    comment_page_number = request.GET.get('comment_page')

    return {
        'participant_page': Paginator(
            _get_participant_queryset(activity, participant_keyword, participant_status), 8
        ).get_page(participant_page_number),
        'participant_keyword': participant_keyword,
        'participant_status': participant_status,
        'participant_querystring': _filtered_get_params(request.GET, ['participant_page']),
        'checkin_page': Paginator(
            _get_checkin_queryset(activity, checkin_keyword, checkin_status), 8
        ).get_page(checkin_page_number),
        'checkin_keyword': checkin_keyword,
        'checkin_status': checkin_status,
        'checkin_querystring': _filtered_get_params(request.GET, ['checkin_page']),
        'moment_page': Paginator(_get_moment_queryset(activity, moment_keyword), 6).get_page(moment_page_number),
        'moment_keyword': moment_keyword,
        'moment_querystring': _filtered_get_params(request.GET, ['moment_page']),
        'comment_page': Paginator(_get_comment_queryset(activity, comment_keyword), 6).get_page(comment_page_number),
        'comment_keyword': comment_keyword,
        'comment_querystring': _filtered_get_params(request.GET, ['comment_page']),
    }


def _build_participant_action_json(activity, registration):
    registrations = ActivityRegistration.objects.filter(activity=activity).select_related('user')
    active_registrations = registrations.exclude(status='cancelled')

    total_registrations = active_registrations.count()
    registered_count = active_registrations.filter(status='registered').count()
    checked_in_count = active_registrations.filter(status='checked_in').count()
    completed_count = active_registrations.filter(status='completed').count()
    cancelled_count = registrations.filter(status='cancelled').count()

    fill_rate = round((total_registrations / activity.max_participants) * 100, 1) if activity.max_participants else 0
    completion_rate = round((completed_count / total_registrations) * 100, 1) if total_registrations else 0

    top_participants_qs = active_registrations.order_by('-user__points', '-registered_at')[:10]
    top_participants = [{
        'username': reg.user.username,
        'department': reg.user.department or '未填写院系',
        'points': reg.user.points,
        'status_display': reg.get_status_display(),
    } for reg in top_participants_qs]

    return {
        'success': True,
        'registration': {
            'id': registration.id,
            'status': registration.status,
            'status_display': registration.get_status_display(),
        },
        'stats': {
            'total_registrations': total_registrations,
            'registered_count': registered_count,
            'checked_in_count': checked_in_count,
            'completed_count': completed_count,
            'cancelled_count': cancelled_count,
            'fill_rate': fill_rate,
            'completion_rate': completion_rate,
        },
        'top_participants': top_participants,
    }


def _build_checkin_action_json(activity, checkin):
    approved_qs = CheckIn.objects.filter(activity=activity, status='approved')
    return {
        'success': True,
        'checkin': {
            'id': checkin.id,
            'status': checkin.status,
            'status_display': checkin.get_status_display(),
            'points_earned': checkin.points_earned,
            'review_note': checkin.review_note or '',
            'reviewer_username': checkin.reviewed_by.username if checkin.reviewed_by else '',
            'revoke_url': reverse('activities:manage_checkin_revoke', args=[activity.id, checkin.id]),
        },
        'stats': {
            'approved': CheckIn.objects.filter(activity=activity, status='approved').count(),
            'pending': CheckIn.objects.filter(activity=activity, status='pending').count(),
            'rejected': CheckIn.objects.filter(activity=activity, status='rejected').count(),
            'revoked': CheckIn.objects.filter(activity=activity, status='revoked').count(),
            'total_points_sent': approved_qs.aggregate(total=Sum('points_earned'))['total'] or 0,
        }
    }


def _build_excel_response(sheet_title, headers, rows, filename, report_title='', filter_lines=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    total_cols = len(headers)
    current_row = 1

    if report_title:
        ws.cell(row=current_row, column=1, value=report_title)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

    for line in filter_lines or []:
        ws.cell(row=current_row, column=1, value=line)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

    if report_title or filter_lines:
        current_row += 1

    header_row = current_row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in rows:
        current_row += 1
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=current_row, column=col_idx, value=value)

    ws.freeze_panes = f'A{header_row + 1}'
    ws.auto_filter.ref = f'A{header_row}:{get_column_letter(total_cols)}{header_row}'

    for col_idx in range(1, total_cols + 1):
        max_length = 12
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, current_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            max_length = max(max_length, len(str(value)) if value is not None else 0)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def activity_list(request):
    now = timezone.now()

    Activity.objects.exclude(status='cancelled').filter(end_time__lt=now).exclude(status='ended').update(status='ended')
    Activity.objects.exclude(status='cancelled').filter(
        start_time__lte=now,
        end_time__gte=now,
    ).exclude(status='ongoing').update(status='ongoing')
    Activity.objects.exclude(status='cancelled').filter(start_time__gt=now).exclude(status='upcoming').update(status='upcoming')

    activities = (
        Activity.objects.all()
        .select_related('category', 'creator')
        .annotate(
            active_registration_count=Count(
                'participants',
                filter=~Q(participants__status='cancelled'),
                distinct=True,
            )
        )
        .prefetch_related('participants')
    )

    q = request.GET.get('q')
    if q:
        activities = activities.filter(Q(title__icontains=q) | Q(location__icontains=q))

    category_id = request.GET.get('category')
    if category_id:
        activities = activities.filter(category_id=category_id)

    status = request.GET.get('status', '')
    if status == 'all':
        pass
    elif status:
        activities = activities.filter(status=status)
    else:
        activities = activities.filter(status__in=['upcoming', 'ongoing'])

    if request.GET.get('available_only'):
        activities = activities.filter(active_registration_count__lt=F('max_participants'), status='upcoming')

    sort = request.GET.get('sort', '-created_at')
    if sort not in {'-created_at', 'created_at', 'start_time', '-start_time', 'view_count', '-view_count'}:
        sort = '-created_at'
    activities = activities.order_by(sort)

    activities = Paginator(activities, 9).get_page(request.GET.get('page'))

    query_params = ''
    if request.GET:
        params = request.GET.copy()
        params.pop('page', None)
        if params:
            query_params = '&' + params.urlencode()

    context = {
        'activities': activities,
        'categories': Category.objects.filter(is_active=True),
        'query_params': query_params,
        'current_status': status,
        'current_category': category_id,
        'current_q': q,
    }
    return render(request, 'activities/list.html', context)


def activity_detail(request, pk):
    activity = get_object_or_404(
        Activity.objects.select_related('category', 'creator').prefetch_related('participants', 'comments', 'managers'),
        pk=pk,
    )

    activity.update_status()
    activity.refresh_from_db()
    activity.view_count += 1
    activity.save(update_fields=['view_count'])

    is_registered = False
    can_checkin = False
    registration = None
    can_edit_activity = False
    can_delete_activity = False
    can_close_activity = False
    can_view_activity_dashboard = False

    if request.user.is_authenticated:
        registration = ActivityRegistration.objects.filter(user=request.user, activity=activity).first()
        is_registered = registration is not None
        if is_registered:
            can_checkin = registration.status == 'registered' and activity.status == 'ongoing'
        can_edit_activity = activity.can_edit(request.user)
        can_delete_activity = activity.can_delete(request.user)
        can_close_activity = activity.can_close(request.user)
        can_view_activity_dashboard = can_edit_activity or can_close_activity

    related_activities = Activity.objects.filter(
        category=activity.category,
        status__in=['upcoming', 'ongoing'],
    ).exclude(id=activity.id)[:3]

    comment_form = ActivityCommentForm()

    activity_dashboard = None
    top_participants = []
    recent_participants = []
    recent_checkins = []
    pending_checkins = []
    rejected_checkins = []
    recent_moments = []
    recent_activity_comments = []
    management_data = {
        'participant_page': None,
        'participant_keyword': '',
        'participant_status': '',
        'participant_querystring': '',
        'checkin_page': None,
        'checkin_keyword': '',
        'checkin_status': '',
        'checkin_querystring': '',
        'moment_page': None,
        'moment_keyword': '',
        'moment_querystring': '',
        'comment_page': None,
        'comment_keyword': '',
        'comment_querystring': '',
    }

    if can_view_activity_dashboard:
        registrations = ActivityRegistration.objects.filter(activity=activity).select_related('user').order_by('-registered_at')
        active_registrations = registrations.exclude(status='cancelled')
        checkins = CheckIn.objects.filter(activity=activity).select_related('user', 'reviewed_by').order_by('-created_at')
        approved_checkins_qs = checkins.filter(status='approved')
        pending_checkins_qs = checkins.filter(status='pending')
        rejected_checkins_qs = checkins.filter(status='rejected')
        revoked_checkins_qs = checkins.filter(status='revoked')

        related_moments = Moment.objects.filter(activity=activity).select_related('user').annotate(
            like_count=Count('likes', distinct=True)
        ).order_by('-created_at')

        total_registrations = active_registrations.count()
        registered_count = active_registrations.filter(status='registered').count()
        checked_in_count = active_registrations.filter(status='checked_in').count()
        completed_count = active_registrations.filter(status='completed').count()
        cancelled_count = registrations.filter(status='cancelled').count()

        approved_count = approved_checkins_qs.count()
        pending_count = pending_checkins_qs.count()
        rejected_count = rejected_checkins_qs.count()
        revoked_count = revoked_checkins_qs.count()

        total_points_sent = approved_checkins_qs.aggregate(total=Sum('points_earned'))['total'] or 0
        total_moments = related_moments.count()
        total_activity_comments = activity.comments.count()
        total_view_count = activity.view_count

        fill_rate = round((total_registrations / activity.max_participants) * 100, 1) if activity.max_participants else 0
        checkin_rate = round((approved_count / total_registrations) * 100, 1) if total_registrations else 0
        completion_rate = round((completed_count / total_registrations) * 100, 1) if total_registrations else 0

        today = timezone.now().date()
        trend_labels = []
        registration_trend_data = []
        checkin_trend_data = []
        for i in range(6, -1, -1):
            date = today - timedelta(days=i)
            trend_labels.append(date.strftime('%m-%d'))
            registration_trend_data.append(registrations.filter(registered_at__date=date).count())
            checkin_trend_data.append(approved_checkins_qs.filter(created_at__date=date).count())

        activity_dashboard = {
            'total_registrations': total_registrations,
            'registered_count': registered_count,
            'checked_in_count': checked_in_count,
            'completed_count': completed_count,
            'cancelled_count': cancelled_count,
            'approved_checkins': approved_count,
            'pending_checkins': pending_count,
            'rejected_checkins': rejected_count,
            'revoked_checkins': revoked_count,
            'total_points_sent': total_points_sent,
            'moment_count': total_moments,
            'activity_comment_count': total_activity_comments,
            'view_count': total_view_count,
            'fill_rate': fill_rate,
            'checkin_rate': checkin_rate,
            'completion_rate': completion_rate,
            'registration_status_labels': ['待参与', '已打卡', '已完成', '已取消'],
            'registration_status_data': [registered_count, checked_in_count, completed_count, cancelled_count],
            'checkin_status_labels': ['通过', '待审核', '拒绝', '已撤销'],
            'checkin_status_data': [approved_count, pending_count, rejected_count, revoked_count],
            'trend_labels': trend_labels,
            'registration_trend_data': registration_trend_data,
            'checkin_trend_data': checkin_trend_data,
        }

        top_participants = active_registrations.order_by('-user__points', '-registered_at')[:10]
        recent_participants = registrations[:8]
        recent_checkins = approved_checkins_qs[:8]
        pending_checkins = pending_checkins_qs[:8]
        rejected_checkins = rejected_checkins_qs[:8]
        recent_moments = related_moments[:6]
        recent_activity_comments = activity.comments.select_related('user').order_by('-created_at')[:6]
        management_data = _build_activity_management_data(activity, request)

    context = {
        'activity': activity,
        'is_registered': is_registered,
        'can_checkin': can_checkin,
        'registration': registration,
        'related_activities': related_activities,
        'comments': activity.comments.filter(parent=None),
        'comment_form': comment_form,
        'can_edit_activity': can_edit_activity,
        'can_delete_activity': can_delete_activity,
        'can_close_activity': can_close_activity,
        'can_view_activity_dashboard': can_view_activity_dashboard,
        'activity_dashboard': activity_dashboard,
        'top_participants': top_participants,
        'recent_participants': recent_participants,
        'recent_checkins': recent_checkins,
        'pending_checkins': pending_checkins,
        'rejected_checkins': rejected_checkins,
        'recent_moments': recent_moments,
        'recent_activity_comments': recent_activity_comments,
        'management_data': management_data,
    }
    return render(request, 'activities/detail.html', context)


@login_required
def activity_create(request):
    messages.info(request, '活动创建需先提交申请；若要发起新的活动，请填写活动申请表。')
    return redirect('activities:application_submit')


@login_required
def activity_edit(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    if not activity.can_edit(request.user):
        messages.error(request, '您没有权限编辑此活动。')
        return redirect('activities:detail', pk=pk)

    if request.method == 'POST':
        form = ActivityForm(request.POST, request.FILES, instance=activity, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, '活动更新成功！')
            return redirect('activities:detail', pk=pk)
        messages.error(request, '更新失败，请检查填写信息。')
    else:
        form = ActivityForm(instance=activity, user=request.user)

    return render(request, 'activities/edit.html', {'form': form, 'categories': Category.objects.filter(is_active=True), 'activity': activity})


@login_required
def activity_delete(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    if not activity.can_delete(request.user):
        messages.error(request, '您没有权限删除此活动。')
        return redirect('activities:detail', pk=pk)

    if request.method == 'POST':
        creator = activity.creator
        activity.delete()
        if creator_id := getattr(creator, 'id', None):
            if request.user.id == creator_id:
                creator.decrement_activity_created()
        messages.success(request, '活动已删除。')
        return redirect('activities:list')

    return render(request, 'activities/delete_confirm.html', {'activity': activity})


@login_required
@require_POST
def close_activity(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    if not activity.can_close(request.user):
        messages.error(request, '只有平台管理员可以关闭活动。')
        return redirect('activities:detail', pk=pk)

    activity.status = 'cancelled'
    activity.closed_by = request.user
    activity.closed_at = timezone.now()
    activity.save(update_fields=['status', 'closed_by', 'closed_at'])

    messages.success(request, '活动已被平台管理员关闭。')
    return redirect('activities:detail', pk=pk)


@login_required
@require_POST
def join_activity(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    activity.update_status()

    if activity.is_full:
        messages.error(request, '活动已满员。')
        return redirect('activities:detail', pk=pk)
    if activity.status != 'upcoming':
        messages.error(request, '活动当前不在报名阶段。')
        return redirect('activities:detail', pk=pk)
    if activity.registration_deadline and timezone.now() > activity.registration_deadline:
        messages.error(request, '报名截止时间已过。')
        return redirect('activities:detail', pk=pk)

    registration, created = ActivityRegistration.objects.get_or_create(
        user=request.user,
        activity=activity,
        defaults={'status': 'registered'},
    )

    if not created and registration.status != 'cancelled':
        messages.warning(request, '您已经报名了此活动。')
        return redirect('activities:detail', pk=pk)

    if not created and registration.status == 'cancelled':
        registration.status = 'registered'
        registration.checked_in_at = None
        registration.save(update_fields=['status', 'checked_in_at'])

    request.user.increment_activity_joined()
    messages.success(request, f'报名成功！本活动完成后可获得 {activity.points} 积分。')
    return redirect('activities:detail', pk=pk)


@login_required
@require_POST
def cancel_registration(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    registration = ActivityRegistration.objects.filter(user=request.user, activity=activity).first()

    if not registration:
        messages.error(request, '您尚未报名此活动。')
        return redirect('activities:detail', pk=pk)
    if registration.status not in ['registered', 'checked_in']:
        messages.error(request, '当前状态无法取消报名。')
        return redirect('activities:detail', pk=pk)

    registration.status = 'cancelled'
    registration.save(update_fields=['status'])
    request.user.decrement_activity_joined()
    messages.success(request, '取消报名成功。')
    return redirect('activities:detail', pk=pk)


@login_required
@require_POST
def add_comment(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    form = ActivityCommentForm(request.POST)

    if form.is_valid():
        comment = form.save(commit=False)
        comment.activity = activity
        comment.user = request.user
        comment.save()
        messages.success(request, '评论发表成功！')
    else:
        messages.error(request, '评论发表失败。')
    return redirect('activities:detail', pk=pk)


@login_required
def my_activities(request):
    activity_type = request.GET.get('type', 'created')

    if activity_type == 'created':
        activities = Activity.objects.filter(creator=request.user)
    elif activity_type == 'joined':
        activities = Activity.objects.filter(
            participants__user=request.user,
            participants__status__in=['registered', 'checked_in', 'completed'],
        ).distinct()
    elif activity_type == 'checked':
        activity_ids = CheckIn.objects.filter(user=request.user).values_list('activity_id', flat=True)
        activities = Activity.objects.filter(id__in=activity_ids)
    elif activity_type == 'managed':
        activities = Activity.objects.filter(managers=request.user).distinct()
    else:
        activities = Activity.objects.filter(creator=request.user)

    return render(request, 'activities/my_activities.html', {
        'activities': activities,
        'current_type': activity_type,
        'user': request.user,
    })


@login_required
def submit_activity_application(request):
    if getattr(request.user, 'role', '') == 'admin':
        messages.error(request, '平台管理员不直接发起活动申请。')
        return redirect('activities:application_list')

    if request.method == 'POST':
        form = ActivityApplicationForm(request.POST, request.FILES)
        if form.is_valid():
            application = form.save(commit=False)
            application.applicant = request.user
            application.save()

            from apps.users.models import User
            from apps.social.models import Message

            platform_admins = User.objects.filter(role='admin', is_active=True)
            for admin_user in platform_admins:
                Message.objects.create(
                    recipient=admin_user,
                    sender=request.user,
                    message_type='system',
                    title='新的活动申请待审核',
                    content=f'用户 {request.user.username} 提交了活动申请《{application.title}》，请尽快审核。',
                )

            messages.success(request, '活动申请已提交，等待平台管理员审核。')
            return redirect('activities:application_list')
        messages.error(request, '活动申请提交失败，请检查表单内容。')
    else:
        form = ActivityApplicationForm()

    return render(request, 'activities/application_form.html', {'form': form})


@login_required
def activity_application_list(request):
    status = request.GET.get('status', '').strip()

    if getattr(request.user, 'role', '') == 'admin':
        applications = ActivityApplication.objects.select_related('applicant', 'category', 'reviewed_by', 'created_activity')
        if status:
            applications = applications.filter(status=status)
    else:
        applications = ActivityApplication.objects.filter(applicant=request.user).select_related('category', 'reviewed_by', 'created_activity')
        if status:
            applications = applications.filter(status=status)

    applications = applications.order_by('-created_at')
    applications_page = Paginator(applications, 10).get_page(request.GET.get('page'))

    return render(request, 'activities/application_list.html', {
        'applications': applications_page,
        'current_status': status,
        'is_platform_admin': getattr(request.user, 'role', '') == 'admin',
    })


@login_required
@require_POST
def approve_activity_application(request, application_id):
    if not _require_platform_admin(request):
        return redirect('activities:application_list')

    application = get_object_or_404(ActivityApplication, pk=application_id)

    if not application.can_be_reviewed:
        messages.warning(request, '该申请已处理，不能重复审核。')
        return redirect('activities:application_list')

    review_note = request.POST.get('review_note', '').strip() or '审核通过'
    with transaction.atomic():
        activity = Activity.objects.create(
            title=application.title,
            description=application.description,
            cover_image=application.cover_image,
            category=application.category,
            creator=application.applicant,
            start_time=application.start_time,
            end_time=application.end_time,
            registration_deadline=application.registration_deadline,
            location=application.location,
            location_lat=application.location_lat,
            location_lng=application.location_lng,
            max_participants=application.max_participants,
            min_participants=application.min_participants,
            points=application.points,
            requirements=application.requirements,
            checkin_radius=application.checkin_radius,
            checkin_review_mode=application.checkin_review_mode,
            status='upcoming',
        )
        activity.update_status()

        applicant = application.applicant
        applicant.increment_activity_created()

        application.status = 'approved'
        application.review_note = review_note
        application.reviewed_by = request.user
        application.reviewed_at = timezone.now()
        application.created_activity = activity
        application.save(
            update_fields=[
                'status', 'review_note', 'reviewed_by',
                'reviewed_at', 'created_activity',
            ]
        )

        Message.objects.create(
            recipient=applicant,
            sender=request.user,
            message_type='system',
            title='活动申请审核通过',
            content=(
                f'你提交的活动申请《{application.title}》已通过审核，系统已自动创建对应活动。'
                '你仅可管理该活动；若后续还想发起新的活动，请重新提交活动申请。'
            ),
            related_activity=activity,
        )

    messages.success(request, '活动申请已通过，系统已自动创建活动。申请人仅获得该活动的管理权限。')
    return redirect('activities:application_list')


@login_required
@require_POST
def reject_activity_application(request, application_id):
    if not _require_platform_admin(request):
        return redirect('activities:application_list')

    application = get_object_or_404(ActivityApplication, pk=application_id)
    if not application.can_be_reviewed:
        messages.warning(request, '该申请已处理，不能重复审核。')
        return redirect('activities:application_list')

    review_note = request.POST.get('review_note', '').strip() or '审核未通过'
    application.status = 'rejected'
    application.review_note = review_note
    application.reviewed_by = request.user
    application.reviewed_at = timezone.now()
    application.save(update_fields=['status', 'review_note', 'reviewed_by', 'reviewed_at'])

    Message.objects.create(
        recipient=application.applicant,
        sender=request.user,
        message_type='system',
        title='活动申请审核未通过',
        content=f'你提交的活动申请《{application.title}》未通过审核。原因：{review_note}',
    )

    messages.success(request, '活动申请已驳回。')
    return redirect('activities:application_list')


@login_required
@require_POST
def manage_registration_cancel(request, pk, registration_id):
    activity = get_object_or_404(Activity, pk=pk)
    if not _can_manage_activity(request.user, activity):
        return _management_permission_denied(request, activity)

    registration = get_object_or_404(ActivityRegistration, pk=registration_id, activity=activity)

    if registration.status == 'cancelled':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': '该报名已是取消状态。'}, status=400)
        messages.warning(request, '该报名已是取消状态。')
        return redirect('activities:detail', pk=pk)

    if registration.status == 'completed':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': '已完成的报名不能直接取消。'}, status=400)
        messages.error(request, '已完成的报名不能直接取消。')
        return redirect('activities:detail', pk=pk)

    with transaction.atomic():
        registration.status = 'cancelled'
        registration.save(update_fields=['status'])
        registration.user.decrement_activity_joined()
        try:
            related_checkin = registration.checkin
        except CheckIn.DoesNotExist:
            related_checkin = None

        if related_checkin and related_checkin.status == 'pending':
            related_checkin.status = 'rejected'
            related_checkin.reviewed_by = request.user
            related_checkin.review_note = '报名已被活动管理员取消'
            related_checkin.save(update_fields=['status', 'reviewed_by', 'review_note'])

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        payload = _build_participant_action_json(activity, registration)
        payload['message'] = f'已取消 {registration.user.username} 的报名。'
        return JsonResponse(payload)

    messages.success(request, f'已取消 {registration.user.username} 的报名。')
    return redirect('activities:detail', pk=pk)


@login_required
@require_POST
def manage_registration_complete(request, pk, registration_id):
    activity = get_object_or_404(Activity, pk=pk)
    if not _can_manage_activity(request.user, activity):
        return _management_permission_denied(request, activity)

    registration = get_object_or_404(ActivityRegistration, pk=registration_id, activity=activity)

    if registration.status == 'completed':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': '该报名已经是完成状态。'}, status=400)
        messages.warning(request, '该报名已经是完成状态。')
        return redirect('activities:detail', pk=pk)

    if registration.status == 'cancelled':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': '已取消的报名不能标记为完成。'}, status=400)
        messages.error(request, '已取消的报名不能标记为完成。')
        return redirect('activities:detail', pk=pk)

    registration.status = 'completed'
    registration.save(update_fields=['status'])

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        payload = _build_participant_action_json(activity, registration)
        payload['message'] = f'已将 {registration.user.username} 标记为完成。'
        return JsonResponse(payload)

    messages.success(request, f'已将 {registration.user.username} 标记为完成。')
    return redirect('activities:detail', pk=pk)


@login_required
@require_POST
def manage_checkin_approve(request, pk, checkin_id):
    activity = get_object_or_404(Activity, pk=pk)
    if not _can_manage_activity(request.user, activity):
        return _management_permission_denied(request, activity)

    checkin = get_object_or_404(CheckIn, pk=checkin_id, activity=activity)

    if checkin.status != 'pending':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': '该打卡不是待审核状态，无法重复处理。'}, status=400)
        messages.warning(request, '该打卡不是待审核状态，无法重复处理。')
        return redirect('activities:detail', pk=pk)

    review_note = request.POST.get('note', '').strip() or '活动管理员审核通过'
    with transaction.atomic():
        checkin.status = 'approved'
        checkin.reviewed_by = request.user
        checkin.review_note = review_note
        checkin.save(update_fields=['status', 'reviewed_by', 'review_note'])

        streak = calculate_continuous_days(checkin.user, checkin.activity)
        points = award_points(checkin.user, activity=activity, streak_days=streak, related_checkin=checkin)

        checkin.points_earned = points
        checkin.save(update_fields=['points_earned'])

        checkin.registration.status = 'completed'
        checkin.registration.save(update_fields=['status'])

        Message.objects.create(
            recipient=checkin.user,
            sender=request.user,
            message_type='activity',
            title='打卡审核已通过',
            content=f'你在活动《{activity.title}》中的打卡已审核通过，获得 {points} 积分。备注：{review_note}',
            related_activity=activity,
        )

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        payload = _build_checkin_action_json(activity, checkin)
        payload['message'] = f'已通过 {checkin.user.username} 的打卡申请。'
        return JsonResponse(payload)

    messages.success(request, f'已通过 {checkin.user.username} 的打卡申请。')
    return redirect('activities:detail', pk=pk)


@login_required
@require_POST
def manage_checkin_reject(request, pk, checkin_id):
    activity = get_object_or_404(Activity, pk=pk)
    if not _can_manage_activity(request.user, activity):
        return _management_permission_denied(request, activity)

    checkin = get_object_or_404(CheckIn, pk=checkin_id, activity=activity)
    if checkin.status != 'pending':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': '该打卡不是待审核状态，无法重复处理。'}, status=400)
        messages.warning(request, '该打卡不是待审核状态，无法重复处理。')
        return redirect('activities:detail', pk=pk)

    review_note = request.POST.get('note', '').strip() or '活动管理员审核拒绝'
    checkin.status = 'rejected'
    checkin.reviewed_by = request.user
    checkin.review_note = review_note
    checkin.save(update_fields=['status', 'reviewed_by', 'review_note'])
    checkin.registration.status = 'registered'
    checkin.registration.save(update_fields=['status'])

    Message.objects.create(
        recipient=checkin.user,
        sender=request.user,
        message_type='activity',
        title='打卡审核未通过',
        content=f'你在活动《{activity.title}》中的打卡未通过审核。原因：{review_note}',
        related_activity=activity,
    )

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        payload = _build_checkin_action_json(activity, checkin)
        payload['message'] = f'已拒绝 {checkin.user.username} 的打卡申请。'
        return JsonResponse(payload)

    messages.success(request, f'已拒绝 {checkin.user.username} 的打卡申请。')
    return redirect('activities:detail', pk=pk)


@login_required
@require_POST
def manage_checkin_revoke(request, pk, checkin_id):
    activity = get_object_or_404(Activity, pk=pk)
    if not _can_manage_activity(request.user, activity):
        return _management_permission_denied(request, activity)

    checkin = get_object_or_404(
        CheckIn.objects.select_related('user', 'registration', 'activity'),
        pk=checkin_id,
        activity=activity,
    )

    if checkin.status != 'approved':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': '只有已通过的打卡才能撤销。'}, status=400)
        messages.warning(request, '只有已通过的打卡才能撤销。')
        return redirect('activities:detail', pk=pk)

    revoke_note = request.POST.get('note', '').strip() or '活动管理员撤销通过'
    with transaction.atomic():
        deducted_points = int(checkin.points_earned or 0)
        if deducted_points > 0:
            checkin.user.add_points(-deducted_points, description=f'撤销活动《{activity.title}》打卡积分：{revoke_note}')

        checkin.status = 'revoked'
        checkin.reviewed_by = request.user
        checkin.review_note = f'撤销通过：{revoke_note}'
        checkin.points_earned = 0
        checkin.save(update_fields=['status', 'reviewed_by', 'review_note', 'points_earned'])

        checkin.registration.status = 'registered'
        checkin.registration.save(update_fields=['status'])

        Message.objects.create(
            recipient=checkin.user,
            sender=request.user,
            message_type='activity',
            title='打卡通过已被撤销',
            content=f'你在活动《{activity.title}》中的已通过打卡已被管理员撤销。原因：{revoke_note}。已回收积分：{deducted_points} 分。',
            related_activity=activity,
        )

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        payload = _build_checkin_action_json(activity, checkin)
        payload['message'] = f'已撤销 {checkin.user.username} 的通过打卡，并回收 {deducted_points} 积分。'
        return JsonResponse(payload)

    messages.success(request, f'已撤销 {checkin.user.username} 的通过打卡，并回收 {deducted_points} 积分。')
    return redirect('activities:detail', pk=pk)


@login_required
@require_POST
def manage_moment_delete(request, pk, moment_id):
    activity = get_object_or_404(Activity, pk=pk)
    if not _can_manage_activity(request.user, activity):
        return _management_permission_denied(request, activity)

    moment = get_object_or_404(Moment.objects.select_related('user'), pk=moment_id, activity=activity)
    owner = moment.user
    preview = (moment.content or '')[:40]
    moment.delete()

    if owner != request.user:
        Message.objects.create(
            recipient=owner,
            sender=request.user,
            message_type='activity',
            title='关联动态已被管理员删除',
            content=f'你在活动《{activity.title}》中的动态已被管理员删除。内容片段：{preview}',
            related_activity=activity,
        )

    remaining_moment_count = Moment.objects.filter(activity=activity).count()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'message': f'已删除 {owner.username} 的关联动态。',
            'moment_id': moment_id,
            'moment_count': remaining_moment_count,
        })

    messages.success(request, f'已删除 {owner.username} 的关联动态。')
    return redirect(f'{reverse("activities:detail", args=[pk])}#social-pane')


@login_required
@require_POST
def manage_activity_comment_delete(request, pk, comment_id):
    activity = get_object_or_404(Activity, pk=pk)
    if not _can_manage_activity(request.user, activity):
        return _management_permission_denied(request, activity)

    comment = get_object_or_404(ActivityComment.objects.select_related('user'), pk=comment_id, activity=activity)
    owner = comment.user
    preview = (comment.content or '')[:40]
    comment.delete()

    if owner != request.user:
        Message.objects.create(
            recipient=owner,
            sender=request.user,
            message_type='activity',
            title='活动评论已被管理员删除',
            content=f'你在活动《{activity.title}》中的评论已被管理员删除。内容片段：{preview}',
            related_activity=activity,
        )

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'message': f'已删除 {owner.username} 的活动评论。',
            'comment_id': comment_id,
            'activity_comment_count': activity.comments.count(),
        })

    messages.success(request, f'已删除 {owner.username} 的活动评论。')
    return redirect(f'{reverse("activities:detail", args=[pk])}#social-pane')


@login_required
def export_activity_participants_csv(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    if not _can_manage_activity(request.user, activity):
        return _export_permission_denied(request, activity)

    qs = _get_participant_queryset(
        activity,
        request.GET.get('participant_q', '').strip(),
        request.GET.get('participant_status', '').strip(),
    )

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="activity_{activity.id}_participants.csv"'
    writer = csv.writer(response)
    writer.writerow(['用户名', '姓名', '院系', '状态', '报名时间'])

    for reg in qs:
        writer.writerow([
            reg.user.username,
            reg.user.real_name,
            reg.user.department,
            reg.get_status_display(),
            reg.registered_at.strftime('%Y-%m-%d %H:%M'),
        ])
    return response


@login_required
def export_activity_checkins_csv(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    if not _can_manage_activity(request.user, activity):
        return _export_permission_denied(request, activity)

    qs = _get_checkin_queryset(
        activity,
        request.GET.get('checkin_q', '').strip(),
        request.GET.get('checkin_status', '').strip(),
    )

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="activity_{activity.id}_checkins.csv"'
    writer = csv.writer(response)
    writer.writerow(['用户名', '状态', '备注', '审核备注', '积分', '打卡时间'])

    for checkin in qs:
        writer.writerow([
            checkin.user.username,
            checkin.get_status_display(),
            checkin.remark,
            checkin.review_note,
            checkin.points_earned,
            checkin.created_at.strftime('%Y-%m-%d %H:%M'),
        ])
    return response


@login_required
def export_activity_moments_csv(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    if not _can_manage_activity(request.user, activity):
        return _export_permission_denied(request, activity)

    qs = _get_moment_queryset(activity, request.GET.get('moment_q', '').strip())

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="activity_{activity.id}_moments.csv"'
    writer = csv.writer(response)
    writer.writerow(['发布者', '内容', '点赞数', '评论数', '图片数', '发布时间'])

    for moment in qs:
        writer.writerow([
            moment.user.username,
            moment.content,
            moment.like_count,
            moment.comment_count,
            moment.images.count(),
            moment.created_at.strftime('%Y-%m-%d %H:%M'),
        ])
    return response


@login_required
def export_activity_participants_excel(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    if not _can_manage_activity(request.user, activity):
        return _export_permission_denied(request, activity)

    keyword = request.GET.get('participant_q', '').strip()
    status = request.GET.get('participant_status', '').strip()
    qs = _get_participant_queryset(activity, keyword, status)

    rows = [[
        reg.user.username,
        reg.user.real_name or '',
        reg.user.department or '',
        reg.get_status_display(),
        reg.registered_at.strftime('%Y-%m-%d %H:%M'),
    ] for reg in qs]

    return _build_excel_response(
        sheet_title='参与者',
        headers=['用户名', '姓名', '院系', '状态', '报名时间'],
        rows=rows,
        filename=f'activity_{activity.id}_participants.xlsx',
        report_title='活动参与者导出',
        filter_lines=[
            f'活动：{activity.title}',
            f'关键词：{keyword or "无"}',
            f'状态：{dict(ActivityRegistration.STATUS_CHOICES).get(status, "全部") if status else "全部"}',
            f'导出时间：{timezone.now().strftime("%Y-%m-%d %H:%M:%S")}',
        ],
    )


@login_required
def export_activity_checkins_excel(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    if not _can_manage_activity(request.user, activity):
        return _export_permission_denied(request, activity)

    keyword = request.GET.get('checkin_q', '').strip()
    status = request.GET.get('checkin_status', '').strip()
    qs = _get_checkin_queryset(activity, keyword, status)

    rows = [[
        checkin.user.username,
        checkin.get_status_display(),
        checkin.remark or '',
        checkin.location_name or '',
        f'{checkin.latitude}, {checkin.longitude}',
        checkin.accuracy,
        checkin.reviewed_by.username if checkin.reviewed_by else '',
        checkin.review_note or '',
        checkin.points_earned,
        checkin.photos.count(),
        checkin.created_at.strftime('%Y-%m-%d %H:%M'),
    ] for checkin in qs]

    return _build_excel_response(
        sheet_title='打卡记录',
        headers=['用户名', '状态', '打卡备注', '位置名称', '经纬度', '定位精度(米)', '审核人', '系统/审核备注', '积分', '照片数', '打卡时间'],
        rows=rows,
        filename=f'activity_{activity.id}_checkins.xlsx',
        report_title='活动打卡记录导出',
        filter_lines=[
            f'活动：{activity.title}',
            f'关键词：{keyword or "无"}',
            f'状态：{dict(CheckIn.STATUS_CHOICES).get(status, "全部") if status else "全部"}',
            f'导出时间：{timezone.now().strftime("%Y-%m-%d %H:%M:%S")}',
        ],
    )


@login_required
def export_activity_moments_excel(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    if not _can_manage_activity(request.user, activity):
        return _export_permission_denied(request, activity)

    keyword = request.GET.get('moment_q', '').strip()
    qs = _get_moment_queryset(activity, keyword)

    rows = [[
        moment.user.username,
        moment.content,
        moment.like_count,
        moment.comment_count,
        moment.images.count(),
        moment.created_at.strftime('%Y-%m-%d %H:%M'),
    ] for moment in qs]

    return _build_excel_response(
        sheet_title='关联动态',
        headers=['发布者', '内容', '点赞数', '评论数', '图片数', '发布时间'],
        rows=rows,
        filename=f'activity_{activity.id}_moments.xlsx',
        report_title='活动关联动态导出',
        filter_lines=[
            f'活动：{activity.title}',
            f'关键词：{keyword or "无"}',
            f'导出时间：{timezone.now().strftime("%Y-%m-%d %H:%M:%S")}',
        ],
    )


class ActivityViewSet(viewsets.ModelViewSet):
    queryset = Activity.objects.filter(status__in=['upcoming', 'ongoing'])
    serializer_class = ActivitySerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        if not self.request.user.can_create_activity():
            raise permissions.PermissionDenied('普通学生不能直接创建活动，请先提交活动申请。')
        serializer.save(creator=self.request.user)


class ActivityRegistrationViewSet(viewsets.ModelViewSet):
    queryset = ActivityRegistration.objects.all()
    serializer_class = ActivityRegistrationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return ActivityRegistration.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
